import pandas as pd
#import nltk
import string
import re
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.cell import Cell
from nltk.corpus import stopwords
#from sklearn.feature_extraction.text import CountVectorizer

df = pd.read_excel('Diku.xlsx', sheet_name='DIKU', usecols='B,C,O,P,Q') #leser DIKU-arket
df.dropna(inplace = True) #skal fjerne null-celler

#kolonner = ['Læringsutbytte - Kunnskap','Læringsutbytte - Ferdigheter','Læringsutbytte - Generell Kompetanse']
keywords = ['digital tvilling','virtuell',' VR[- ]',' AR[- ]',' XR[- ]','hololens','big room','revit','programvare','trimble'
,' BIM[- ]','digital samhand','digital','modell','kunstlig intelligens',' ICE[- ]',' VDC[- ]','samtidig prosjektering'
,' IPD[- ]','lean', 'maskinlæring',' AI[- ]',' IFC[- ]','maker','samarbeid','teknologi','studentaktiv','problembasert','programm','script'] #søkeord

#string.punctuation uten bindestrek
tegnsetting = """!"#$%&'()*+,./:;<=>?@[\]^_`{|}~'"""

#metode som fjerner tegnsetting, utenom bindestrek, og stopwords
def text_process(frame): 
    nopunc = [char for char in frame if char not in tegnsetting]
    nopunc = ''.join(nopunc)   
    nopunc = [word for word in nopunc.split() if word.lower not in stopwords.words('norwegian')]
    return nopunc

#tokenized og prosessert versjon av læringsutbytte kolonnene
LUK = df['Læringsutbytte - Kunnskap'].apply(text_process)
LUF = df['Læringsutbytte - Ferdigheter'].apply(text_process)
LUG = df['Læringsutbytte - Generell Kompetanse'].apply(text_process)


wb = load_workbook(filename='resultat.xlsx') #output excel ark
ws = wb['Statistikk'] #velger sheet
wb.save(filename='resultat.xlsx') #lagrer

#variabler som blir brukt i output
Emnekode = df['Emnekode']
Emnenavn = df['Emnenavn']

#antall fag
max_row = len(df)

#variabler som blir brukt i senere metoder
count_max = 0 #teller for hele søkeordet
actual_max = 0 #teller for alle søkeord

#lager md-fil til alternativ output
md = open("resultat.md", "w+")


def search_in_frame(frame, k):
    i = 0 #indeks for celle
    count = 0 #teller for et enkelt frame
    global count_max
    global actual_max
    for _ in frame:
        
        #bow_transformer = CountVectorizer(analyzer=text_process).fit(frame[i])
        #unique += (len(bow_transformer.vocabulary_))
        arraystr = ' '.join(map(str, frame[i])) #setter sammen igjen meldingen for printing 
        search = re.search(keywords[k].lower(),arraystr.lower()) #søkefunskjon
        if str(search) != 'None': #sjekker at det er match                                  
            
            #workaround for sheetnavn hvor [- ] var med i søkeordet
            title = keywords[k]
            if "[- ]" in title:
                    title = title.replace('[- ]','')
            
            for sheet in wb.worksheets: #sjekker om sheet med match allerede finnes
                createsheet = True
                if sheet.title == title:
                    createsheet = False
                    break
                else:
                    continue

            if createsheet == True: #lager ny sheet hvis den ikke fantes               
                ws2 = wb.create_sheet(title)
                ws2.cell(1,1, 'LUK:')
                ws2.cell(2,1, 'Emnekode:')
                ws2.cell(2,2,'Emnenavn:')
                ws2.cell(2,3, 'Læringsutbytte')
                ws2.cell(1,4, 'LUF:')
                ws2.cell(2,4, 'Emnekode:')
                ws2.cell(2,5,'Emnenavn:')
                ws2.cell(2,6, 'Læringsutbytte')
                ws2.cell(1,7, 'LUG:')
                ws2.cell(2,7, 'Emnekode:')
                ws2.cell(2,8,'Emnenavn:')
                ws2.cell(2,9, 'Læringsutbytte')
            else: #hvis den fantes så skriver den i riktig ark
                ws2 = wb[title]

            #skriver til nye sheets
            if str(frame) == str(LUK): #LUK info
                ws2.cell(count+3,1, Emnekode[i])
                ws2.cell(count+3,2, Emnenavn[i])
                ws2.cell(count+3,3, arraystr)
            elif str(frame) == str(LUF):#LUF info
                ws2.cell(count+3,4, Emnekode[i])
                ws2.cell(count+3,5, Emnenavn[i])
                ws2.cell(count+3,6, arraystr)
            else:#LUG info
                ws2.cell(count+3,7, Emnekode[i])
                ws2.cell(count+3,8, Emnenavn[i])
                ws2.cell(count+3,9, arraystr)

            for row in ws2.iter_rows(): #setter alle celler til bryt tekst
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
            
            #setter widt slik at den er mer leselig
            ws2.column_dimensions['A'].width = 15
            ws2.column_dimensions['D'].width = 15
            ws2.column_dimensions['G'].width = 15

            ws2.column_dimensions['B'].width = 20
            ws2.column_dimensions['E'].width = 20
            ws2.column_dimensions['H'].width = 20

            ws2.column_dimensions['C'].width = 50
            ws2.column_dimensions['F'].width = 50
            ws2.column_dimensions['I'].width = 50

            print(Emnekode[i]+': '+arraystr+'\n') #printer ut emnekode og meldingen
            md.write(Emnekode[i]+': '+arraystr+'\n\n') #skriver det samme til resultat.md
                  
            count = count + 1 #+1 i teller
            i = i + 1 #neste celle
            continue
        i = i + 1 #neste celle hvis det ikke vvar treff
        
    print(str(count)+' treff av '+str(max_row)+' mulige\n') #printer ut antall treff
    md.write(str(count)+' treff av '+str(max_row)+' mulige\n')
    count_max += count #legger til dette i max-count for keyword
    
    if str(frame) == str(LUK):
        ws.cell(k+2,2,count) #skriver til excel ark
    elif str(frame) == str(LUF):
       ws.cell(k+2,3,count)
    else: #sjekker at det er siste frame
        print(str(count_max)+' treff av totalt '+str((max_row*3))+' mulige på søkeordet: '+keywords[k]) #printer ut max_count
        md.write(str(count_max)+' treff av totalt '+str((max_row*3))+' mulige på søkeordet: '+keywords[k]+'\n')
        ws.cell(k+2,4,count) #skriver til statistikk arket
        ws.cell(k+2,5,count_max)#skriver til statistikk arket
        actual_max += count_max #legger til count_max til actual_max
        count_max = 0 #resetter count_max
       
def wordsearch(k):
    p = 0   #indeks for frame
    for _ in range(3):
        if p == 0: #går første gjennom LUK
            print('LUK:')
            md.write('LUK: \n')
            search_in_frame(LUK, k) #kaller search_in_fram funksjonen i LUK
        elif p == 1: #så LUF
            print('LUF:')
            md.write('LUF: \n')
            search_in_frame(LUF, k) #kaller search_in_fram funksjonen i LUF
        else: #til sist LUG
            print('LUG:')
            md.write('LUG: \n')
            search_in_frame(LUG, k) #kaller search_in_fram funksjonen i LUG
        p = p + 1 #neste frame

def main():

    for sheet in wb.worksheets: #sletter alle sheets bortsett fra første
        if sheet.title == 'Statistikk':
            continue
        else:
            wb.remove(sheet)

    k = 0 #indeks for keyword

    for _ in keywords: #rydding i output
        if keywords[k].startswith(' '):
            x = keywords[k].replace(' ','',1)
            print('\n'+x+':')
            md.write('\n'+x+':\n')
            ws.cell(k+2,1,x)
        else:
            print('\n'+keywords[k]+':')
            md.write('\n'+keywords[k]+':\n')
            ws.cell(k+2,1,keywords[k])
            
        
        wordsearch(k) #kjører søk
        k = k + 1 #neste søkeord

    max_mulige = (max_row*3) * len(keywords) #max mulige antall treff

    print('\n\n'+str(actual_max)+' treff av totalt '+str(max_mulige)+' mulige.') #printer totale treff
    md.write('\n\n'+str(actual_max)+' treff av totalt '+str(max_mulige)+' mulige.')#printer totale treff
    ws.cell(2,7,actual_max) #skriver antall totale treff
    ws.cell(2,9,max_row) #antall mulige treff for LUK
    ws.cell(2,10,max_row) #antall mulige treff for LUF
    ws.cell(2,11,max_row) #antall mulige treff for LUG
    ws.cell(2,12,max_row*3) #antall mulige treff per søkeord
    ws.cell(2,13,max_mulige)#antall mulige treff totalt

    for _ in range(ws.max_row): #fjerner ord fra excel-arket som lenger ikke er i keywords
        if ws.cell(1,len(keywords)+2) != None:
            ws.delete_rows(len(keywords)+2)
            
    wb.save(filename='resultat.xlsx') #lagrer excel fil



main() #gjør alt basically
