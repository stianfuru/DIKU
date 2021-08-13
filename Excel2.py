import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from nltk.corpus import stopwords


df = pd.read_excel('Diku.xlsx', sheet_name='DIKU', usecols='B,C,L,M,N')
df.dropna(inplace=True)


keywords = ['prosjektbasert','prosjekt','gruppe','studentaktiv','flipped']

tegnsetting = """!"#$%&'()*+,./:;<=>?@[\]^_`{|}~'"""

def text_process(frame): 
    nopunc = [char for char in frame if char not in tegnsetting]
    nopunc = ''.join(nopunc)   
    nopunc = [word for word in nopunc.split() if word.lower not in stopwords.words('norwegian')]
    return nopunc


LOA = df['Læringsformer og aktiviteter'].apply(text_process)
AK = df['Arbeidskrav'].apply(text_process)
EF = df['Eksamensform'].apply(text_process)

wb = load_workbook(filename='arbeidskrav.xlsx')
ws = wb['Statistikk']
wb.save(filename='arbeidskrav.xlsx')

Emnekode = df['Emnekode']
Emnenavn = df['Emnenavn']

max_row = len(df)

count_max = 0
actual_max = 0
unique = []



def search_in_frame(frame, k):
    i = 0 
    count = 0
    global count_max
    global actual_max
    for _ in frame:

        arraystr = ' '.join(map(str, frame[i])) #setter sammen igjen meldingen for printing og søk
        search = re.search(keywords[k].lower(),arraystr.lower()) #søkefunskjon
        if str(search) != 'None':
            
            title = keywords[k]

            for sheet in wb.worksheets:
                createsheet = True
                if sheet.title == title:
                    createsheet = False
                    break
                else:
                    continue

            if createsheet == True:
                ws2 = wb.create_sheet(title)
                ws2.cell(1,1, 'LOA')
                ws2.cell(2,1, 'Emnekode:')
                ws2.cell(2,2, 'Emnenavn:')
                ws2.cell(2,3, 'Læringsformer og aktiviteter:')
                ws2.cell(1,4, 'AK')
                ws2.cell(2,4, 'Emnekode:')
                ws2.cell(2,5, 'Emnenavn:')
                ws2.cell(2,6, 'Arbeidskrav')
                ws2.cell(1,7, 'EF')
                ws2.cell(2,7, 'Emnekode:')
                ws2.cell(2,8, 'Emnenavn:')
                ws2.cell(2,9, 'Eksamensform')
            else:
                ws2 = wb[title]

            if str(frame) == str(LOA): #LUK info
                ws2.cell(count+3,1, Emnekode[i])
                ws2.cell(count+3,2, Emnenavn[i])
                ws2.cell(count+3,3, arraystr)
            elif str(frame) == str(AK):#LUF info
                ws2.cell(count+3,4, Emnekode[i])
                ws2.cell(count+3,5, Emnenavn[i])
                ws2.cell(count+3,6, arraystr)
            else:#LUG info
                ws2.cell(count+3,7, Emnekode[i])
                ws2.cell(count+3,8, Emnenavn[i])
                ws2.cell(count+3,9, arraystr)

            for row in ws2.iter_rows(): #setter alle celler til bryt tekst
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
            
            ws2.column_dimensions['A'].width = 15
            ws2.column_dimensions['D'].width = 15
            ws2.column_dimensions['G'].width = 15

            ws2.column_dimensions['B'].width = 20
            ws2.column_dimensions['E'].width = 20
            ws2.column_dimensions['H'].width = 20

            ws2.column_dimensions['C'].width = 50
            ws2.column_dimensions['F'].width = 50
            ws2.column_dimensions['I'].width = 50

            count = count + 1 #+1 i teller
            i = i + 1 #neste celle
            continue
        i = i + 1 #neste celle hvis det ikke var treff

    count_max += count

    if str(frame) == str(LOA):
        ws.cell(k+2,2,count) 
    elif str(frame) == str(AK):
       ws.cell(k+2,3,count)
    else: 
        ws.cell(k+2,4,count)
        ws.cell(k+2,5,count_max)
        actual_max += count_max
        count_max = 0

def wordsearch(k):
    p = 0
    for _ in range(3):
        if p == 0:
            search_in_frame(LOA, k)
        elif p == 1:
            search_in_frame(AK, k)
        else:
            search_in_frame(EF, k)
        p = p + 1
def main():

    for sheet in wb.worksheets:
        if sheet.title == 'Statistikk':
            continue
        else:
            wb.remove(sheet)

    k = 0

    for _ in keywords:
        ws.cell(k+2,1,keywords[k])
        wordsearch(k)
        k = k + 1

    max_mulige = (max_row*3) * len(keywords)
    ws.cell(2,7,actual_max) #skriver antall totale treff
    ws.cell(2,10,max_row) #antall mulige treff per kategori
    ws.cell(2,11,max_row*3) #antall mulige treff per søkeord
    ws.cell(2,12,max_mulige)#antall mulige treff totalt

    for _ in range(ws.max_row): #fjerner ord fra statistikk-arket som lenger ikke er i keywords
        if ws.cell(1,len(keywords)+2) != None:
            ws.delete_rows(len(keywords)+2)
    
    wb.save(filename='arbeidskrav.xlsx')


main()